from __future__ import annotations

import datetime, json, pathlib, random, shutil, sys, re, os
import tkinter as tk
from   tkinter import filedialog, messagebox

import numpy as np, pandas as pd, openpyxl
from   openpyxl.utils import get_column_letter
from   openpyxl.styles import PatternFill                       # NEW

# ─────────────────────────────────────────────────────────────────────────────
#  Seat catalogue – mirrors the visual grid
# ─────────────────────────────────────────────────────────────────────────────
SEATS: dict[str, dict[str, object]] = {

    # ---------- Workstations -------------------------------------------------
    **{f"WS {n}": {"type": "ws",
                   "adjustable": n == 6,
                   "seat_number": n}
       for n in range(1, 11)},                       # 1-10  (WS-6 adjustable)

    # ---------- Regular open-area seats -------------------------------------
    **{f"Seat {n}": {"type": "reg",
                     "adjustable": n in (13, 15, 30),
                     "seat_number": n}
       for n in range(1, 31)},                       # 1-30, 13/15/30 adjustable

    # ---------- Private rooms on campus (main building) ---------------------
    **{f"Room {n}": {"type": "pr",
                     "adjustable": n in (345, 347, 348, 349, 351, 354),
                     "seat_number": n}
       for n in range(345, 355)},                    # Rooms 345-354

    # ---------- *Extra* private rooms at Campus Corners ---------------------
    **{f"CC Room {n}": {"type": "pr",
                        "adjustable": False,         # none are height-adj
                        "seat_number": n}
       for n in range(1, 26)},                       # 25 rooms labelled CC 1-25

    # ---------- SAS testing offices -----------------------------------------
    **{f"SAS {n}": {"type": "sas",
                    "adjustable": n == 5,
                    "seat_number": n}
       for n in range(1, 13)},                       # 1-12, SAS-5 adjustable

    # ---------- SHA classrooms ----------------------------------------------
    # SHA-356 : 17 seats (seat 1 adjustable)
    **{f"SHA 356 Seat {n}": {"type": "sha",
                             "classroom": 356,
                             "adjustable": n == 1,
                             "seat_number": n}
       for n in range(1, 18)},

    # SHA-357 : 19 seats (seat 1 adjustable)
    **{f"SHA 357 Seat {n}": {"type": "sha",
                             "classroom": 357,
                             "adjustable": n == 1,
                             "seat_number": n}
       for n in range(1, 20)},

    # SHA-358 : 31 seats (seats 1-3 adjustable)
    **{f"SHA 358 Seat {n}": {"type": "sha",
                             "classroom": 358,
                             "adjustable": n in (1, 2, 3),
                             "seat_number": n}
       for n in range(1, 32)},

    # SHA-359 : **15** seats (seat 1 adjustable)   ← fixed count
    **{f"SHA 359 Seat {n}": {"type": "sha",
                             "classroom": 359,
                             "adjustable": n == 1,
                             "seat_number": n}
       for n in range(1, 16)},
}

# ───────────────────────────── Constants ──────────────────────────────────
TIME_FORMAT_EXCEL = "h:mm AM/PM"      # makes 22:00 look like 10:00 PM
COLOR_BY_TYPE   = {                   # SHA=yellow, SAS=blue, …
    "sha": "FFF2CC",
    "sas": "DDEBF7",
    "pr":  "E2F0D9",
    "ws":  "FCE4D6",
    "reg": None,                      # no colour for regular seats
}
MIN_COL_WIDTH = 10

# ───────────────────────────── GUI helpers ──────────────────────────────────
def pick_file(title: str, patterns: tuple[tuple[str, str], ...]) -> str:
    """Choose a file via Tk when possible, otherwise fall back to CLI input."""
    try:
        root = tk.Tk(); root.withdraw()
        path = filedialog.askopenfilename(title=title, filetypes=patterns)
        root.destroy()
        return path
    except tk.TclError:
        # Headless environment – ask in the terminal
        return input(f"{title}: ")

def pick_folder(title: str) -> str:
    """Choose a folder via Tk when possible, otherwise fall back to CLI."""
    try:
        root = tk.Tk(); root.withdraw()
        folder = filedialog.askdirectory(title=title)
        root.destroy()
        return folder
    except tk.TclError:
        return input(f"{title}: ")


def ask_yes_no(title: str, message: str, default: bool = False) -> bool:
    """Display a yes/no dialog or fall back to terminal input."""
    try:
        root = tk.Tk(); root.withdraw()
        result = messagebox.askyesno(title, message)
        root.destroy()
        return bool(result)
    except tk.TclError:
        prompt = f"{title}: {message} [{'Y/n' if default else 'y/N'}]: "
        resp = input(prompt)
        if resp.strip() == '':
            return default
        return resp.strip().lower() in ('y', 'yes')


def choose_room_preferences() -> dict[str, bool]:
    """Ask the user which room types to include.

    Uses a Tk checkbox window when possible. In headless environments,
    falls back to simple yes/no prompts in the terminal.
    """
    try:
        root = tk.Tk()
        root.title("Select Room Types")
        vars_ = {
            'private_rooms':   tk.BooleanVar(value=True),
            'campus_corners':  tk.BooleanVar(value=True),
            'workstations':    tk.BooleanVar(value=True),
            'regular_seats':   tk.BooleanVar(value=True),
            'sas_offices':     tk.BooleanVar(value=True),
            'sha_classrooms':  tk.BooleanVar(value=True),
        }
        for text, var in [
            ("Private Rooms (Main Bldg)", vars_['private_rooms']),
            ("Campus Corners Rooms",      vars_['campus_corners']),
            ("Workstations",              vars_['workstations']),
            ("Regular Seats",             vars_['regular_seats']),
            ("SAS Offices",               vars_['sas_offices']),
            ("SHA Classrooms",            vars_['sha_classrooms']),
        ]:
            tk.Checkbutton(root, text=text, variable=var).pack(anchor="w")

        prefs: dict[str, bool] = {}

        def _ok() -> None:
            for key, var in vars_.items():
                prefs[key] = var.get()
            root.destroy()

        tk.Button(root, text="OK", command=_ok).pack(pady=5)
        root.mainloop()
        return prefs
    except tk.TclError:
        # Terminal fallback
        defaults = {
            'private_rooms':   True,
            'campus_corners':  True,
            'workstations':    True,
            'regular_seats':   True,
            'sas_offices':     True,
            'sha_classrooms':  True,
        }
        prefs: dict[str, bool] = {}
        for key, default in defaults.items():
            prompt = f"Include {key.replace('_', ' ')}? [{'Y/n' if default else 'y/N'}]: "
            resp = input(prompt)
            if resp.strip() == '':
                prefs[key] = default
            else:
                prefs[key] = resp.strip().lower() in ('y', 'yes')
        return prefs
# ───────────────────────────── Data ingest ──────────────────────────────────
def read_source(path: str) -> pd.DataFrame:
    """Load roster, tag 'Requires Adjustable', coerce time-like columns."""
    df = (pd.read_csv(path, header=1)
          if path.lower().endswith(".csv")
          else pd.read_excel(path, header=1))

    df["Requires Adjustable"] = (
        df["Test Accommodation"]
        .str.contains("Height Adjustable", case=False, na=False)
    )

    def to_time(col: str) -> pd.Series:
        # Suppress the dateutil warning by being more specific
        with pd.option_context('mode.chained_assignment', None):
            return (pd.to_datetime(df[col].astype(str), errors="coerce", format='mixed')
                    .dt.time.fillna(datetime.time.min))

    for col in ("Begin Time", "End Time", "Class Time"):
        df[col] = to_time(col)

    return df

# ─────────────────────── Seat-assignment engine ─────────────────────────────
def assign_students(df: pd.DataFrame, seat_pool: list[str]
                    ) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Greedy, stable assignment of students to seats without overlapping times:
    adjustable-desk requests are placed first; within each sub-group students
    are processed in chronological order.  The pool is shuffled every pass
    for fairness.
    """
    availability = {s: datetime.time.min for s in seat_pool}
    rng          = np.random.default_rng()
    placed, left = [], []

    for needs_adj in (True, False):
        group = df[df["Requires Adjustable"] == needs_adj]\
                  .sort_values("Begin Time")
        for _, stu in group.iterrows():
            valid = [s for s in seat_pool
                     if (not needs_adj or SEATS[s]["adjustable"])]
            rng.shuffle(valid)

            for seat in valid:
                if stu["Begin Time"] >= availability[seat]:
                    availability[seat] = stu["End Time"]
                    placed.append({**stu, "Test Room": seat})
                    break
            else:
                left.append(stu)

    return pd.DataFrame(placed), pd.DataFrame(left)

# ───────────────────────────── Constants and helpers ──────────────────────────────────
def _set_cell(ws, row, col, value, *, fill_hex=None):
    """
    Write *value* into (row, col), prettify if it's a time,
    and optionally apply a background fill colour.
    """
    cell = ws.cell(row=row, column=col, value=value)

    # Make times human-readable in Excel
    if isinstance(value, datetime.time):
        cell.number_format = TIME_FORMAT_EXCEL

    # Shade the cell if a colour was supplied
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)

    return cell

# ───────────────────────────── Enhanced Excel output helper ──────────────────────────────
def write_excel(df: pd.DataFrame, name: str,
                template: str | None,
                out_dir: str | pathlib.Path,
                assignment_type: str = "ASSIGNED") -> None:

    out_path = pathlib.Path(out_dir) / f"{name}.xlsx"

    # ── 1  Set up the workbook ────────────────────────────────────────────────
    if template:
        shutil.copy(template, out_path)
        wb = openpyxl.load_workbook(out_path)
        # Try to use "Master" sheet, fall back to active sheet if not found
        try:
            ws = wb["Master"]
        except KeyError:
            ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Create basic headers if no template
        headers = ['Begin Time', 'End Time', 'Student Number', 'Student Last Name', 'Student First Name',
                  'Check-IN Time', 'Check-OUT Time', 'Course', 'Code', 'Test Room', 'Seat Number', 
                  'Faculty Name', 'Class Time', 'Test Accommodation', 'Invigilator Comment', 'Test Comment']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col_idx, value=header)

    # ── 2  Locate or create "Test Room" & "Seat Number" columns ───────────────
    hdr_row      = 2
    headers      = {ws.cell(hdr_row, col).value: col for col in range(1, ws.max_column+1)}
    test_room_col   = headers.get("Test Room")  or ws.max_column + 1
    seat_number_col = headers.get("Seat Number") or (test_room_col + 1)

    ws.cell(hdr_row, test_room_col,   "Test Room")
    ws.cell(hdr_row, seat_number_col, "Seat Number")

    # ── 3  Clear previous data completely ────────────────────────────────────
    if ws.max_row > hdr_row:
        for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
            for cell in row:
                cell.value = None   # wipe values & formula results

    # ── 4  Write this cohort's data ───────────────────────────────────────────
    for excel_row, (_, stu) in enumerate(df.iterrows(), start=hdr_row + 1):

        # Core student columns
        _set_cell(ws, excel_row, 1,  stu.get('Begin Time'))          # Begin Time
        _set_cell(ws, excel_row, 2,  stu.get('End Time'))            # End Time
        _set_cell(ws, excel_row, 3,  stu.get('Student Number'))
        _set_cell(ws, excel_row, 4,  stu.get('Student Last Name'))
        _set_cell(ws, excel_row, 5,  stu.get('Student First Name'))
        _set_cell(ws, excel_row, 8,  stu.get('Course'))
        _set_cell(ws, excel_row, 9,  stu.get('Code'))
        _set_cell(ws, excel_row, 12, stu.get('Faculty Name'))
        _set_cell(ws, excel_row, 13, stu.get('Class Time'))
        _set_cell(ws, excel_row, 14, stu.get('Test Accommodation'))

        # ----------  Test Room & Seat Number  ----------
        if assignment_type == "ASSIGNED" and pd.notna(stu.get('Test Room')):
            test_room = stu['Test Room']

            # Pick colour based on room type
            seat_type = SEATS.get(test_room, {}).get("type", "reg")
            fill_hex  = COLOR_BY_TYPE.get(seat_type)

            _set_cell(ws, excel_row, test_room_col, test_room, fill_hex=fill_hex)

            # Seat number logic unchanged
            seat_no = (SEATS.get(test_room, {}).get("seat_number")
                       or re.search(r"\b(?:Seat|WS|Room|SAS) (\d+)", test_room))
            if seat_no:
                _set_cell(ws, excel_row, seat_number_col,
                          seat_no.group(1) if hasattr(seat_no, "group") else seat_no)

    # ── 5  Adjust widths, save, done ─────────────────────────────────────────
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        if all(cell.value is None for cell in column[2:]):
            ws.column_dimensions[letter].width = MIN_COL_WIDTH
            continue
        max_len = max(len(str(cell.value)) for cell in column if cell.value)
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    wb.save(out_path)
    print(f"📄  Saved {out_path.name}  ({len(df)} rows)")

# ─────────────────────────────── Main ───────────────────────────────────────
def main() -> None:
    # 1  Pick the roster
    data = pick_file("Select your Excel/CSV roster",
                     (("Excel/CSV", "*.xlsx *.xls *.csv"),))
    if not data:
        print("No data file selected – exiting."); sys.exit()

    # 2  Ask whether Excel outputs are wanted
    want_excel = ask_yes_no(
        "Excel outputs?",
        "Generate the usual cohort workbooks as well?\n"
        "(No → skip straight to JSON output only.)",
        default=True,
    )
    
    # 3  Pick template (default to May 5, 2025.xlsx)
    template = None
    if want_excel:
        use_default_template = ask_yes_no(
            "Use Default Template?",
            "Use 'May 5, 2025.xlsx' as the template?\n\n"
            "This will maintain the standard format with proper\n"
            "Test Room and Seat Number columns.\n\n"
            "Choose 'No' to select a different template.",
            default=True,
        )
        
        if use_default_template:
            template_path = pathlib.Path(__file__).parent / "May 5, 2025.xlsx"
            if template_path.exists():
                template = str(template_path)
                print(f"📋 Using default template: {template}")
            else:
                print("❌ May 5, 2025.xlsx not found in current directory")
                template = pick_file("Select an .xlsx template",
                                   (("Excel", "*.xlsx *.xls"),))
        else:
            template = pick_file("Select an .xlsx template",
                               (("Excel", "*.xlsx *.xls"),))
        
        if not template:
            print("No template chosen; Excel outputs will be basic workbooks.")

    # 4  Select room types
    room_preferences = choose_room_preferences()

    # 5  Pick an output folder
    out = pick_folder("Select an output folder")
    if not out:
        print("No output folder – exiting."); sys.exit()

    # 6  Load & preprocess the roster
    df = read_source(data)

    # 7  Cohort splits - Sequential building to avoid overlaps
    print("🔍 Building cohorts sequentially...")
    print(f"📋 Total students in roster: {len(df)}")
    
    # Start with all students
    remaining = df.copy()
    
    # Sequential cohort building
    print("\n🔄 Building cohorts sequentially...")
    
    # DF1_PR - Private Room (highest priority)
    pr_mask = remaining["Test Accommodation"].str.contains("Private Room", case=False, na=False)
    DF1_PR = remaining[pr_mask]
    remaining = remaining.drop(DF1_PR.index)
    print(f"   DF1_PR: {len(DF1_PR)} students (remaining: {len(remaining)})")
    
    # DF2_WS - Workstation needs
    ws_mask = remaining["Test Accommodation"].str.contains(r"Read and Write|MS Word|Kurzweil", case=False, na=False)
    DF2_WS = remaining[ws_mask]
    remaining = remaining.drop(DF2_WS.index)
    print(f"   DF2_WS: {len(DF2_WS)} students (remaining: {len(remaining)})")
    
    # DF3_HAD - Height Adjustable Desks
    DF3_HAD = remaining[remaining["Requires Adjustable"]]
    remaining = remaining.drop(DF3_HAD.index)
    print(f"   DF3_HAD: {len(DF3_HAD)} students (remaining: {len(remaining)})")
    
    # DF4_SCRIBE - Scribe needs
    scribe_mask = remaining["Test Accommodation"].str.contains("Scribe", case=False, na=False)
    DF4_SCRIBE = remaining[scribe_mask]
    remaining = remaining.drop(DF4_SCRIBE.index)
    print(f"   DF4_SCRIBE: {len(DF4_SCRIBE)} students (remaining: {len(remaining)})")
    
    # DF6_FINAL - Final exams
    final_mask = remaining["Test Accommodation"].str.contains("Final", case=False, na=False)
    DF6_FINAL = remaining[final_mask]
    remaining = remaining.drop(DF6_FINAL.index)
    print(f"   DF6_FINAL: {len(DF6_FINAL)} students (remaining: {len(remaining)})")
    
    # DF7_ES - Evening Students
    end_time_22 = remaining["End Time"].apply(lambda t: t.hour == 22 if hasattr(t, 'hour') else False)
    begin_before_class = remaining["Begin Time"] < remaining["Class Time"]
    es_mask = end_time_22 & begin_before_class
    DF7_ES = remaining[es_mask]
    remaining = remaining.drop(DF7_ES.index)
    print(f"   DF7_ES: {len(DF7_ES)} students (remaining: {len(remaining)})")
    
    # ---------- SAS offices (must run BEFORE we create MAIN) --------------------
    if room_preferences['sas_offices']:
        sas_mask = remaining["Test Accommodation"].str.contains(
            r"SAS|Special|Individual|Separate|Extra Time|Alternative|Modified",
            case=False, na=False
        )
        DF8_SAS = remaining[sas_mask]
        remaining = remaining.drop(DF8_SAS.index)
        print(f"   DF8_SAS: {len(DF8_SAS)} students (remaining: {len(remaining)})")
    else:
        DF8_SAS = pd.DataFrame()
        print(f"   DF8_SAS: {len(DF8_SAS)} students (SAS offices disabled)")

    # ---------- DF5_MAIN – whatever is still unhandled --------------------------
    DF5_MAIN = remaining.copy()
    remaining = remaining.drop(DF5_MAIN.index)
    print(f"   DF5_MAIN: {len(DF5_MAIN)} students (remaining: {len(remaining)})")

    # ---------- DF9_CLASSROOMS – absolute leftovers -----------------------------
    DF9_CLASSROOMS = remaining.copy()
    print(f"   DF9_CLASSROOMS: {len(DF9_CLASSROOMS)} students")
    
    # Verify no overlaps
    total_in_cohorts = len(DF1_PR) + len(DF2_WS) + len(DF3_HAD) + len(DF4_SCRIBE) + len(DF5_MAIN) + len(DF6_FINAL) + len(DF7_ES) + len(DF8_SAS) + len(DF9_CLASSROOMS)
    print(f"\n📊 Cohort verification:")
    print(f"   Total students in cohorts: {total_in_cohorts}")
    print(f"   Original roster size: {len(df)}")
    print(f"   ✅ No overlaps: {total_in_cohorts == len(df)}")

    splits = {
        "DF1_PR": DF1_PR, "DF2_WS": DF2_WS, "DF3_HAD": DF3_HAD,
        "DF4_SCRIBE": DF4_SCRIBE, "DF5_MAIN": DF5_MAIN,
        "DF6_FINAL": DF6_FINAL, "DF7_ES": DF7_ES, "DF8_SAS": DF8_SAS, "DF9_CLASSROOMS": DF9_CLASSROOMS
    }
    
    if want_excel:
        print("📊 Generating cohort analysis files...")
        for name, part in splits.items():
            write_excel(part, name, template, out, "PROCESSED")

    # 8  Build seat pools based on user preferences
    seat_pools = {}
    
    # --- Private rooms (main building) -----------------------------------------
    if room_preferences['private_rooms']:
        seat_pools['private_rooms'] = [s for s in SEATS
                                       if SEATS[s]["type"] == "pr"
                                       and not s.startswith("CC Room")]
        print(f"✅ Main-building private rooms: {len(seat_pools['private_rooms'])} seats")
    else:
        seat_pools['private_rooms'] = []

    # --- Campus Corners --------------------------------------------------------
    if room_preferences.get('campus_corners', False):
        cc_rooms = [s for s in SEATS if s.startswith("CC Room")]
        seat_pools['private_rooms'].extend(cc_rooms)      # piggy-back on same pool
        print(f"✅ Campus Corners rooms added: {len(cc_rooms)} seats")
    
    if room_preferences['workstations']:
        seat_pools['workstations'] = [s for s in SEATS if SEATS[s]["type"] == "ws"]
        print(f"✅ Workstations enabled: {len(seat_pools['workstations'])} seats")
    
    if room_preferences['regular_seats']:
        seat_pools['regular_seats'] = [s for s in SEATS if SEATS[s]["type"] == "reg"]
        print(f"✅ Regular seats enabled: {len(seat_pools['regular_seats'])} seats")
    
    if room_preferences['sas_offices']:
        seat_pools['sas_offices'] = [s for s in SEATS if SEATS[s]["type"] == "sas"]
        print(f"✅ SAS offices enabled: {len(seat_pools['sas_offices'])} seats")

    if room_preferences['sha_classrooms']:
        seat_pools['sha_classrooms'] = [s for s in SEATS if SEATS[s]["type"] == "sha"]
        print(f"✅ SHA classrooms enabled: {len(seat_pools['sha_classrooms'])} seats")

    # Collect all adjustable seats from the enabled pools
    adjustable_sources = []
    for pool_name in ("private_rooms", "workstations", "regular_seats",
                      "sas_offices", "sha_classrooms"):
        adjustable_sources.extend(seat_pools.get(pool_name, []))
    seat_pools['adjustable_seats'] = [s for s in adjustable_sources
                                      if SEATS[s]["adjustable"]]
    if seat_pools['adjustable_seats']:
        print(f"✅ Adjustable seats available: {len(seat_pools['adjustable_seats'])} seats")

    # 9  Seat-assignment cohorts
    cohorts = {}

    # Assign cohorts to appropriate seat pools
    if room_preferences['private_rooms'] and len(DF1_PR) > 0:
        cohorts["DF1_PR"] = (DF1_PR, seat_pools['private_rooms'])

    if len(seat_pools.get('adjustable_seats', [])) > 0 and len(DF3_HAD) > 0:
        cohorts["DF3_HAD"] = (DF3_HAD, seat_pools['adjustable_seats'])

    if room_preferences['private_rooms'] and len(DF4_SCRIBE) > 0:
        cohorts["DF4_SCRIBE"] = (DF4_SCRIBE, seat_pools['private_rooms'])

    if room_preferences['workstations'] and len(DF2_WS) > 0:
        cohorts["DF2_WS"] = (DF2_WS, seat_pools['workstations'])

    if room_preferences['sha_classrooms'] and len(DF6_FINAL) > 0:
        cohorts["DF6_FINAL"] = (DF6_FINAL, seat_pools['sha_classrooms'])

    if room_preferences['sha_classrooms'] and len(DF7_ES) > 0:
        cohorts["DF7_ES"] = (DF7_ES, seat_pools['sha_classrooms'])

    if room_preferences['regular_seats'] and len(DF5_MAIN) > 0:
        cohorts["DF5_MAIN"] = (DF5_MAIN, seat_pools['regular_seats'])
    
    if room_preferences['sas_offices'] and len(DF8_SAS) > 0:
        cohorts["DF8_SAS"] = (DF8_SAS, seat_pools['sas_offices'])
    
    if room_preferences['sha_classrooms'] and len(DF9_CLASSROOMS) > 0:
        cohorts["DF9_CLASSROOMS"] = (DF9_CLASSROOMS, seat_pools['sha_classrooms'])

    # Ensure every cohort – even empty ones – has its three workbooks
    if want_excel:
        for name, part in splits.items():
            if (out_path := pathlib.Path(out, f"{name}.xlsx")).exists():
                continue                         # raw workbook already written
            write_excel(part, name,                template, out, "PROCESSED")
            write_excel(part, f"{name}_ASSIGNED",  template, out, "ASSIGNED")
            write_excel(part, f"{name}_NOT_ASSIGNED", template, out, "NOT_ASSIGNED")

    # 10  Process assignments with proper duplicate counting
    assigns: dict[str, dict[str, object]] = {}
    seen = set()  # Track student numbers already placed
    assigned_ids = set()  # Track all assigned student IDs
    not_assigned_ids = set()  # Track all not-assigned student IDs
    
    for name, (part, pool) in cohorts.items():
        print(f"🔄 Processing {name}: {len(part)} students, {len(pool)} seats available")
        assigned, not_assigned = assign_students(part, pool)
        
        # Guard against double-seating
        assigned_unique = assigned[~assigned["Student Number"].isin(seen)]
        double_seated = assigned[assigned["Student Number"].isin(seen)]
        
        if len(double_seated) > 0:
            print(f"   ⚠️  Prevented double-seating: {len(double_seated)} students")
            # Move double-seated students to not_assigned
            not_assigned = pd.concat([not_assigned, double_seated], ignore_index=True)
        
        # Update tracking
        seen.update(assigned_unique["Student Number"])
        assigned = assigned_unique  # Use only unique assignments
        
        # Track unique student IDs for proper counting (with safety checks)
        if not assigned.empty:
            assigned_ids.update(assigned["Student Number"])
        if not not_assigned.empty:
            not_assigned_ids.update(not_assigned["Student Number"])
        
        print(f"   ✅ Assigned: {len(assigned)}, ❌ Not assigned: {len(not_assigned)}")

        if want_excel:
            write_excel(assigned,     f"{name}_ASSIGNED",     template, out, "ASSIGNED")
            write_excel(not_assigned, f"{name}_NOT_ASSIGNED", template, out, "NOT_ASSIGNED")

        # build JSON payload (with safety check)
        if not assigned.empty:
            for _, row in assigned.iterrows():
                assigns[row["Test Room"]] = {
                    "student_number": int(row["Student Number"]),
                    "last_name":      str(row["Student Last Name"]),
                    "first_name":     str(row["Student First Name"]),
                    "requiresAdjust": bool(row["Requires Adjustable"]),
                }

    # Handle empty cohorts - create empty ASSIGNED and NOT_ASSIGNED files for missing cohorts
    empty_cohorts = set(splits) - set(cohorts)
    for name in empty_cohorts:
        print(f"📝 Creating empty files for cohort: {name}")
        if want_excel:
            write_excel(pd.DataFrame(), f"{name}_ASSIGNED",     template, out, "ASSIGNED")
            write_excel(pd.DataFrame(), f"{name}_NOT_ASSIGNED", template, out, "NOT_ASSIGNED")

    # 11  Save JSON files and provide detailed summary
    pathlib.Path(out, "seats.json").write_text(
        json.dumps(SEATS, indent=2), encoding="utf-8")
    pathlib.Path(out, "assigns.json").write_text(
        json.dumps(assigns, indent=2), encoding="utf-8")

    # Calculate final totals using unique student counts
    total_assigned = len(assigned_ids)
    total_not_assigned = len(not_assigned_ids)

    # Print detailed completion summary
    print(f"\n✅  All done! Excel reports and JSON files generated in:")
    print(f"    📁 {out}")
    print(f"\n📊 Assignment Summary:")
    print(f"    • Total students assigned: {total_assigned}")
    print(f"    • Total students not assigned: {total_not_assigned}")
    print(f"    • Total students processed: {total_assigned + total_not_assigned}")
    print(f"    • Original roster size: {len(df)}")
    if total_assigned + total_not_assigned > 0:
        print(f"    • Assignment success rate: {(total_assigned/(total_assigned+total_not_assigned)*100):.1f}%")
    
    # Verify totals match
    if total_assigned + total_not_assigned != len(df):
        print(f"    ⚠️  Warning: Total processed ({total_assigned + total_not_assigned}) != roster size ({len(df)})")
    else:
        print(f"    ✅ Totals verified: All {len(df)} students accounted for")

    # ────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()

# ─────────────────────────────── Debug helper ───────────────────────────────────────
def debug_dataframe_splits(df: pd.DataFrame) -> None:
    """Standalone function to debug why certain cohorts are empty."""
    print("🔍 DEBUGGING DATAFRAME SPLITS")
    print("=" * 50)
    
    print(f"📋 Total students in roster: {len(df)}")
    print(f"📋 Column names: {list(df.columns)}")
    
    # Check if required columns exist
    required_columns = ["Test Accommodation", "Begin Time", "End Time", "Class Time"]
    for col in required_columns:
        if col in df.columns:
            print(f"✅ '{col}' column exists")
        else:
            print(f"❌ '{col}' column MISSING!")
    
    if "Test Accommodation" in df.columns:
        print(f"\n🔍 Test Accommodation analysis:")
        print(f"   Non-null accommodations: {df['Test Accommodation'].notna().sum()}")
        print(f"   Null accommodations: {df['Test Accommodation'].isna().sum()}")
        
        # Show sample accommodations
        sample_accommodations = df["Test Accommodation"].dropna().head(10)
        print(f"   Sample accommodations:")
        for i, acc in enumerate(sample_accommodations):
            print(f"     {i+1}. '{acc}'")
    
    # Check Requires Adjustable column
    if "Requires Adjustable" in df.columns:
        print(f"\n🔍 'Requires Adjustable' column:")
        print(f"   True: {df['Requires Adjustable'].sum()}")
        print(f"   False: {(~df['Requires Adjustable']).sum()}")
        print(f"   Null: {df['Requires Adjustable'].isna().sum()}")
    else:
        print(f"\n❌ 'Requires Adjustable' column not found!")
        
    # Test time columns
    time_columns = ["Begin Time", "End Time", "Class Time"]
    for col in time_columns:
        if col in df.columns:
            print(f"\n🔍 '{col}' column:")
            print(f"   Type: {df[col].dtype}")
            print(f"   Non-null: {df[col].notna().sum()}")
            print(f"   Sample values: {df[col].dropna().head(3).tolist()}")
    
    print("\n" + "=" * 50)
